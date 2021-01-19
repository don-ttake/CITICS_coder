"""
Author: Mingchen Li
Occupation: CITICS Investment
Email:lmc9781@outlook.com
Abstract:
    this is some script that dedicates for beginner to warm up using python and openpyxl
    Sophisticated version will be ExcelHandlePro
"""
import xlrd
#xlrd不支持xlsx了，只支持xls

import openpyxl
# 给出excel文件绝对路径
loc = ("nasdaq100.xlsx")
# 打开工作表
wb = openpyxl.load_workbook(loc)

#获取工作簿 workbook的所有工作表
shenames=wb.sheetnames
print(shenames)

# 这里读取的是第一个sheet
sheet = wb.worksheets[0]
print("第一张表的title:")
print(sheet.title)

print("第一张表的行数和列数:")
rows=sheet.max_row
columns=sheet.max_column
print(rows,columns)

for cell in list(sheet.rows)[3]:  #获取第四行的数据
    print(cell.value,end=" ")
print()

#输出特定的列
for cell in list(sheet.columns)[2]:  #获取第三列的数据
    print(cell.value,end=" ")
print()

