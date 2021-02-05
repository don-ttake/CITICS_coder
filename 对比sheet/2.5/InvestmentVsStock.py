"""
Author: Mingchen Li
Occupation: CITICS Investment
Email:lmc9781@outlook.com
Abstract:
    this is some script that dedicates for news feed for comparing two csv sheets based on investment entities
"""

import csv
import os
import difflib
from openpyxl import Workbook


def equalPrice(a, b):
    """
    自己定义的equal price，主要还是得input的时候strip一下空格才会准， 原稿中0.00和‘-’的问题目前只想到这个方法做
    :param a:
    :param b:
    :return:
    """
    if (a == '0.00' and b == '-'):
        return True
    if (b == '0.00' and a == '-'):
        return True
    return (str(a) == str(b))


MAX_CHAR = 26


def longestSubstringFinder(string1, string2):
    """
    找出最长mutual substring， 中英文都好使
    :param string1:
    :param string2:
    :return:
    """
    answer = ""
    len1, len2 = len(string1), len(string2)
    for i in range(len1):
        match = ""
        for j in range(len2):
            if (i + j < len1 and string1[i + j] == string2[j]):
                match += string2[j]
            else:
                if (len(match) > len(answer)): answer = match
                match = ""
    if (answer in ['上海', '金石', '科技', '生物', '股份', '公司','广州','江苏']):
        return ""
    return answer


print(longestSubstringFinder('0106棒谷科技', '棒谷网络'))


def buildAB(path='./'):
    """
    读path下面的对比A对比B Excel，以B为准，找到名字一样的就写到output excel（B表.xlsx）里，
    然后找钱一样但是不是0的，然后空一个写到output excel（B表.xlsx）里
    最后没找到的B空一行写到后面
    :param path:
    :return:
    """
    loc = os.path.join(path, "中证底稿.csv")
    Anames = []
    Aprices = []
    with open(loc) as csv_file:
        csv_reader = csv.reader(csv_file)
        linecountA = 0
        for row in csv_reader:
            Anames.append(row[0].strip())
            Aprices.append(row[1].strip())
            linecountA += 1
    loc = os.path.join(path, "库务底稿.csv")
    Bnames = []
    Bprices = []
    with open(loc) as csv_file:
        csv_reader = csv.reader(csv_file)
        linecountB = 0
        for row in csv_reader:
            Bnames.append(row[0].strip())
            Bprices.append(row[1].strip())
            linecountB += 1
    found_counter = 0
    sameNames = []
    allSameNames = []
    sameNamesPrices = []
    allSameNamesPrices = []
    pload0 = []
    pload0.append(['AB表存在共同名字的投资'])

    for i in range(0, linecountA):
        for j in range(0, linecountB):
            if ((Anames[i] in Bnames[j]) or
                    (Bnames[j] in Anames[i]) or
                    len(longestSubstringFinder(Anames[i], Bnames[j])) > 1 or
                    len(longestSubstringFinder(Bnames[j], Anames[i])) > 1):
                print("发现对应名字", Anames[i], Bnames[j])
                print("测试投资金额", equalPrice(Aprices[i], Bprices[j]), Aprices[i], Bprices[j])
                sameNames.append(min([Anames[i], Bnames[j]], key=len))
                sameNamesPrices.append(max([Aprices[i], Bprices[j]], key=len))
                allSameNames.append(Anames[i])
                allSameNames.append(Bnames[j])
                allSameNamesPrices.append(Aprices[i])
                allSameNamesPrices.append(Bprices[j])
                pload0.append([Anames[i], Aprices[i], Bnames[j], Bprices[j], equalPrice(Aprices[i], Bprices[j])])
                found_counter += 1
    # Bnames_stripped=Bnames.remove(allSameNames)
    # Bnames_stripped = Bprices.remove(allSameNamesPrices)

    print("对应数有", found_counter, " A 表有", linecountA, " B 表有", linecountB)
    # writer(sameNames, sameNamesPrices, 'sameNames')
    # print(sameNamesPrices , sameNames)
    samePriceNames = []
    samePriceNamesPrice = []
    samePriceCounter = 0
    pload1 = []
    pload1.append(['AB表存在同样投资额，但是名字不同的'])

    for i in range(0, linecountA):
        for j in range(0, linecountB):
            if (equalPrice(Aprices[i], Bprices[j]) and Aprices[i] not in ['0.00', '-']):
                if (Anames[i] not in allSameNames and Bnames not in allSameNames):
                    print("发现对应金额名字", Anames[i], Bnames[j])
                    print("测试投资金额", equalPrice(Aprices[i], Bprices[j]), Aprices[i], Bprices[j])
                    samePriceNames.append(min([Anames[i], Bnames[j]], key=len))
                    samePriceNamesPrice.append(max([Aprices[i], Bprices[j]], key=len))
                    pload1.append([Anames[i], Aprices[i], Bnames[j], Bprices[j], equalPrice(Aprices[i], Bprices[j])])
                    allSameNames.append(Anames[i])
                    allSameNames.append(Bnames[j])
                    allSameNamesPrices.append(Aprices[i])
                    allSameNamesPrices.append(Bprices[j])
                    samePriceCounter += 1

    print("对应数有", samePriceCounter)
    # writer(samePriceNames, samePriceNamesPrice,'samePrices')
    pload2 = []
    pload2.append(['A表B表剩下的'])
    for i in range(0, linecountB):
        if (Bnames[i] not in allSameNames):
            pload2.append(['', '', Bnames[i], Bprices[i]])
    for i in range(0, linecountA):
        if (Anames[i] not in allSameNames):
            pload2.append([Anames[i], Aprices[i], '', ''])
    poladWriter([pload0, pload1, pload2], 'B表')
    print('' in Anames)
    print(Anames)


def poladWriter(ploadList, filename):
    dest = filename + '.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['(A)中证表企业名称', '中证表对应投资金额', '(B)库务表企业名称', '库务表对应投资金额', '是否金额相等'])
    for pload in ploadList:
        for row in pload:
            worksheet.append(row)
        worksheet.append(['', '', '', ''])
    workbook.save(dest)


def writer(names, prices, filename):
    dest = filename + '.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    row = 1
    col = 1

    # Iterate over the data and write it out row by row.
    for i in range(0, len(names)):
        worksheet.cell(row=row, column=col, value=names[i])
        worksheet.cell(row=row, column=col + 1, value=prices[i])
        row += 1

    # # Write a total using a formula.
    # worksheet.write(row, 0, 'Total')
    # worksheet.write(row, 1, '=SUM(B1:B4)')

    workbook.save(dest)


buildAB()
